"""End-to-end tests against the real bundled Master DB spreadsheet.

Mirrors ``test_ground_motion_catalog.py``'s style: load the actual shipped
data file (not a synthetic fixture) and assert on what it is really supposed
to contain, so a future edit to the spreadsheet that breaks these
expectations is caught here.
"""

from pathlib import Path

import pytest

from openframe.core.domain import (
    PropertyResolutionStatus,
    RecordNotFoundError,
    density_kg_m3_to_unit_weight_kN_m3,
)
from openframe.infrastructure.material_section_db import (
    load_database_from_excel,
    load_material_database,
)

_EXCEL_PATH = (
    Path(__file__).resolve().parents[2]
    / "src"
    / "openframe"
    / "infrastructure"
    / "material_section_db"
    / "data"
    / "material_section_database.xlsx"
)


@pytest.fixture(scope="module")
def db():
    return load_material_database(force_reload=True)


def test_bundled_excel_file_exists() -> None:
    assert _EXCEL_PATH.exists()


def test_loading_the_real_workbook_finds_every_sheet_worth_of_records(db) -> None:
    assert len(db.all_materials()) == 43
    # 8 original seed sections (1 per shape) + 26 standard KS D 3502 H-beam
    # sizes added on top of the single H/I Section seed row.
    assert len(db.all_sections()) == 34
    assert len(db.all_rc_sections()) == 3
    assert len(db.all_property_rules()) == 10
    assert len(db.all_sources()) == 15


def test_no_duplicate_ids_across_materials_sections_rc_sections_or_sources(db) -> None:
    assert len({m.material_id for m in db.all_materials()}) == len(db.all_materials())
    assert len({s.section_id for s in db.all_sections()}) == len(db.all_sections())
    assert len({rc.section_id for rc in db.all_rc_sections()}) == len(db.all_rc_sections())
    assert len({s.source_id for s in db.all_sources()}) == len(db.all_sources())


# -- Required lookups --------------------------------------------------------


def test_c24_concrete_lookup(db) -> None:
    c24 = db.get_material("CONC-C24")

    assert c24.category == "Concrete"
    assert c24.fck_MPa == 24
    assert c24.E_MPa == pytest.approx(25811.00626094313)
    assert c24.auto_appliable is True  # VERIFIED + auto_apply=True


def test_sm355_structural_steel_lookup_has_a_pending_conditional_yield_strength(db) -> None:
    sm355 = db.get_material("STL-SM355")

    assert sm355.category == "Structural Steel"
    assert sm355.E_MPa == 200000
    assert sm355.fy_MPa is None  # left to PropertyRules, not filled in directly
    assert sm355.auto_appliable is False  # PENDING

    resolution = db.resolve_property("STL-SM355", "fy_MPa", context={"thickness_mm": 10})
    # A rule interval exists for SM355 at this thickness, but its property_value
    # cell is empty in the Master DB - must be reported, never invented.
    assert resolution.status is PropertyResolutionStatus.RULE_VALUE_MISSING
    assert resolution.value is None


def test_sd400_rebar_lookup(db) -> None:
    sd400 = db.get_material("REBAR-SD400")

    assert sd400.category == "Reinforcing Steel"
    assert sd400.fy_MPa == 400
    assert sd400.auto_appliable is False  # PENDING despite having a concrete number


def test_mdf_general_lookup_resolves_elastic_modulus_by_thickness_rule(db) -> None:
    mdf = db.get_material("MDF-GENERAL")
    assert mdf.density_kg_m3 == 750

    resolved = db.resolve_property("MDF-GENERAL", "E_MPa", context={"thickness_mm": 15})
    assert resolved.status is PropertyResolutionStatus.RESOLVED_BY_RULE
    assert resolved.value == 2200
    assert resolved.auto_appliable is False  # PENDING

    missing_context = db.resolve_property("MDF-GENERAL", "E_MPa")
    assert missing_context.status is PropertyResolutionStatus.MISSING_CONTEXT


def test_user_defined_material_is_handled_as_a_placeholder_not_a_zero(db) -> None:
    user_defined = db.get_material("USER-DEFINED")
    assert user_defined.E_MPa is None

    resolution = db.resolve_property("USER-DEFINED", "E_MPa")
    assert resolution.status is PropertyResolutionStatus.USER_DEFINED
    assert resolution.auto_appliable is False


def test_unknown_material_id_raises_instead_of_returning_none(db) -> None:
    with pytest.raises(RecordNotFoundError):
        db.get_material("NOT-A-REAL-MATERIAL-ID")


def test_pending_materials_are_blocked_from_automatic_application(db) -> None:
    pending_materials = [m for m in db.all_materials() if m.data_status.value == "PENDING"]
    assert len(pending_materials) == 31  # matches the spreadsheet's own VAL-011 notice
    assert all(material.auto_appliable is False for material in pending_materials)
    # And the concrete grades are the only ones actually cleared for auto-apply.
    auto_appliable_ids = {m.material_id for m in db.all_materials() if m.auto_appliable}
    assert auto_appliable_ids == {
        "CONC-C18", "CONC-C21", "CONC-C24", "CONC-C27", "CONC-C30",
        "CONC-C35", "CONC-C40", "CONC-C50", "CONC-C60", "CONC-C70", "CONC-C80",
    }


def test_property_rules_conditional_lookup_across_the_full_thickness_range(db) -> None:
    below_range = db.resolve_property("STL-SS275", "fy_MPa", context={"thickness_mm": 0})
    first_band = db.resolve_property("STL-SS275", "fy_MPa", context={"thickness_mm": 16})
    second_band = db.resolve_property("STL-SS275", "fy_MPa", context={"thickness_mm": 40})
    beyond_range = db.resolve_property("STL-SS275", "fy_MPa", context={"thickness_mm": 41})

    assert below_range.status is PropertyResolutionStatus.NO_MATCHING_RULE  # 0 is excluded
    assert (first_band.status, first_band.value) == (PropertyResolutionStatus.RESOLVED_BY_RULE, 275)
    assert second_band.status is PropertyResolutionStatus.RULE_VALUE_MISSING  # rule exists, empty
    assert beyond_range.status is PropertyResolutionStatus.NO_MATCHING_RULE


def test_unit_conversion_matches_every_material_that_declares_both_fields(db) -> None:
    checked = 0
    for material in db.all_materials():
        if material.density_kg_m3 is None or material.unit_weight_kN_m3 is None:
            continue
        expected = density_kg_m3_to_unit_weight_kN_m3(material.density_kg_m3)
        assert material.unit_weight_kN_m3 == pytest.approx(expected, rel=1e-6), material.material_id
        checked += 1
    assert checked > 0


def test_sections_and_materials_stay_independent(db) -> None:
    h_section = db.get_section("SEC-H-300X300X10X15")
    assert not hasattr(h_section, "E_MPa")
    assert not hasattr(h_section, "material_id")
    # Geometry from the section, material properties looked up separately.
    steel = db.get_material("STL-SM275")
    assert h_section.area_mm2 == 11700
    assert steel.E_MPa == 200000


def test_rc_section_combines_concrete_rebar_geometry_and_layout(db) -> None:
    rc = db.get_rc_section("RC-RECT-300X500-C24-SD400")
    concrete, rebar = db.get_rc_section_materials("RC-RECT-300X500-C24-SD400")

    assert concrete.material_id == "CONC-C24"
    assert rebar.material_id == "REBAR-SD400"
    assert rc.width_mm == 300
    assert rc.height_mm == 500
    assert rc.longitudinal_bar_count == 4
    assert rc.cover_mm == 40


# -- Cache round-trip --------------------------------------------------------


def test_cache_round_trip_matches_a_fresh_excel_parse(tmp_path) -> None:
    cache_path = tmp_path / "cache.json"

    fresh = load_database_from_excel(_EXCEL_PATH)
    from_excel = load_material_database(
        excel_path=_EXCEL_PATH, cache_path=cache_path, force_reload=True
    )
    assert cache_path.exists()
    from_cache = load_material_database(
        excel_path=_EXCEL_PATH, cache_path=cache_path, force_reload=False
    )

    assert {m.material_id for m in from_cache.all_materials()} == {
        m.material_id for m in fresh.all_materials()
    }
    c24_from_cache = from_cache.get_material("CONC-C24")
    c24_fresh = fresh.get_material("CONC-C24")
    assert c24_from_cache == c24_fresh
    assert from_excel.warnings == from_cache.warnings


def test_cache_is_invalidated_when_the_source_file_changes(tmp_path) -> None:
    import openpyxl

    excel_copy = tmp_path / "copy.xlsx"
    excel_copy.write_bytes(_EXCEL_PATH.read_bytes())
    cache_path = tmp_path / "cache.json"

    first_load = load_material_database(excel_path=excel_copy, cache_path=cache_path, force_reload=True)
    assert first_load.get_material("CONC-C24").note is not None
    original_cache_text = cache_path.read_text(encoding="utf-8")

    workbook = openpyxl.load_workbook(excel_copy)
    worksheet = workbook["Materials"]
    header = [cell.value for cell in worksheet[1]]
    id_column = header.index("material_id") + 1
    note_column = header.index("note") + 1
    target_row = next(
        row for row in range(2, worksheet.max_row + 1)
        if worksheet.cell(row=row, column=id_column).value == "CONC-C24"
    )
    worksheet.cell(row=target_row, column=note_column, value="EDITED FOR CACHE INVALIDATION TEST")
    workbook.save(excel_copy)

    second_load = load_material_database(excel_path=excel_copy, cache_path=cache_path, force_reload=False)

    assert second_load.get_material("CONC-C24").note == "EDITED FOR CACHE INVALIDATION TEST"
    assert cache_path.read_text(encoding="utf-8") != original_cache_text
